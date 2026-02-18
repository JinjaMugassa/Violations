"""Pull GALCO _ ALL EVENTS_V1 report and build EVENTS/SUMMARY output.

Uses group template 115 to discover unit IDs, then pulls single-unit template 117
to capture Event text. For each unit and event type, keeps the first occurrence
and sets Count of Event time to total occurrences.
"""

import os
import sys
import json
import time
import html
import copy
import argparse
from datetime import datetime, timezone, timedelta

import pandas as pd
import requests
import openpyxl
import pytz

from wialon_api import WialonAPI
from utils import get_local_timezone_offset, find_column


DEFAULT_GROUP = "TRANSIT_ALL_TRUCKS"
GROUP_TEMPLATE_ID = 115  # GALCO _ ALL EVENTS_V1 (Group)
UNIT_TEMPLATE_ID = 117   # GALCO _ ALL EVENTS_V1 (Single Unit)
DEFAULT_OUTPUT_DIR = r"C:\Users\SAMA\Downloads\GALCO_ALL_EVENTS"
DEFAULT_SAMPLE_NAME = "TRANSIT_ALL_TRUCKS_EVENTS_15.12.2025.xlsx"

DEFAULT_ACTION = "Posted by Traking officer"
DEFAULT_REMARK = "Shared report to the risk group"

SUMMARY_REMARKS = {
    "HARSH BRAKING": "Reported to the risk group",
    "IDLING": "Reported to the risk group",
    "MAIN POWER DISCONNECTED": (
        "I have followed up and checked the battery status, and all of them are okay. "
        "Additionally, many trucks are here in the Kurasini yard, at the borders, and in the mines."
    ),
    "PARKED MORE THAN 3HRS": (
        "All the trucks that have been parked for over 3 hours have been reported in the risk group on WhatsApp"
    ),
    "SPEEDING": (
        "Escalated TC, I contacted the drivers who exceeded the speed limit of 81 and reminded them to move within the speed limit."
    ),
    "NIGHT DRIVING": "Reported to the risk group",
}

SUMMARY_ORDER = [
    "HARSH BRAKING",
    "IDLING",
    "MAIN POWER DISCONNECTED",
    "PARKED MORE THAN 3HRS",
    "NIGHT DRIVING",
    "SPEEDING",
]

EXCLUDED_EVENT_TYPES = {"NIGHT DRIVING", "LOW SPEED"}
TANZANIA_TZ = pytz.timezone("Africa/Dar_es_Salaam")


def parse_date_arg(date_str):
    if not date_str:
        return None
    s = str(date_str).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    raise ValueError("Invalid date format. Use YYYY-MM-DD or DD.MM.YYYY")


def get_system_today_date():
    """Return today's date in Tanzania timezone."""
    return datetime.now(TANZANIA_TZ).date()


def get_day_interval_tz(target_date, until_now=False):
    tz_local = timezone(timedelta(hours=3))
    start_dt = datetime(target_date.year, target_date.month, target_date.day, 0, 0, 0, tzinfo=tz_local)
    if until_now:
        now_local = datetime.now(tz_local)
        end_dt = now_local if now_local >= start_dt else start_dt
    else:
        end_dt = start_dt + timedelta(days=1) - timedelta(seconds=1)
    return int(start_dt.timestamp()), int(end_dt.timestamp())


def format_ddmmyyyy_hhmmss(val):
    try:
        if pd.isna(val) or str(val).strip() == "":
            return ""
        dt = pd.to_datetime(val, dayfirst=True, errors="coerce")
        if pd.isna(dt):
            return str(val).strip()
        # Wialon report datetime values are treated as UTC and displayed in Tanzania time.
        if getattr(dt, "tzinfo", None) is None:
            dt = pytz.UTC.localize(dt.to_pydatetime()).astimezone(TANZANIA_TZ)
        else:
            dt = dt.tz_convert(TANZANIA_TZ)
        return dt.strftime("%d.%m.%Y %H:%M:%S")
    except Exception:
        return str(val).strip()


def infer_notification_text(text):
    if not text:
        return ""
    t = str(text).lower()
    if "parked for over 3hrs" in t or "parked more than 3hrs" in t or "parked for over 3" in t:
        return "PARKED MORE THAN 3HRS"
    if "low speed" in t:
        return "LOW SPEED"
    if "harsh" in t and "brak" in t:
        return "HARSH BRAKING"
    if "idling" in t:
        return "IDLING"
    if "main power" in t and "disconnect" in t:
        return "MAIN POWER DISCONNECTED"
    if "night driving" in t:
        return "NIGHT DRIVING"
    if "speeding" in t or "over speed" in t:
        return "SPEEDING"
    if "speed" in t:
        return "SPEEDING"
    return ""


def should_exclude_event(event_text, notification_text):
    txt = str(event_text or "").lower()
    nt = str(notification_text or "").strip().upper()
    if nt in EXCLUDED_EVENT_TYPES:
        return True
    if "night driving" in txt:
        return True
    if "low speed" in txt:
        return True
    return False


def exec_report_raw(api, object_id, template_id, interval_from, interval_to, output_path):
    tz_offset = get_local_timezone_offset()
    params = {
        "svc": "report/exec_report",
        "params": json.dumps({
            "reportResourceId": api.RESOURCE_ID,
            "reportTemplateId": int(template_id),
            "reportObjectId": int(object_id),
            "reportObjectSecId": 0,
            "interval": {"from": int(interval_from), "to": int(interval_to), "flags": 0},
            "tzOffset": tz_offset,
        }),
        "sid": api.sid,
    }

    resp = requests.post(api.api_url, params=params, timeout=30)
    data = resp.json()
    if "reportResult" not in data:
        print("Report execution failed:", data)
        return None, None, None

    tables = data["reportResult"].get("tables", [])
    if not tables:
        print("No tables in report result")
        return None, None, data

    table = tables[0]
    headers = table.get("header", [])
    row_count = table.get("rows", 0)

    rows_list = []
    if row_count > 0:
        rows_list = api._fetch_report_rows(row_count, output_path)

    return headers, rows_list, data


def save_unit_debug_json(output_path, unit_id, unit_name, headers, rows_list):
    """Save per-unit pulled rows for traceability."""
    try:
        base_name = os.path.splitext(os.path.basename(output_path))[0]
        debug_dir = os.path.join(os.path.dirname(output_path), f"{base_name}_unit_debug")
        os.makedirs(debug_dir, exist_ok=True)
        safe_name = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(unit_name))
        debug_path = os.path.join(debug_dir, f"unit_{unit_id}_{safe_name}.json")
        payload = {
            "unit_id": unit_id,
            "unit_name": unit_name,
            "headers": headers or [],
            "rows": rows_list or [],
        }
        with open(debug_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def save_group_debug_json(output_path, headers, rows_list):
    """Save the group pull rows for traceability."""
    try:
        base_name = os.path.splitext(os.path.basename(output_path))[0]
        debug_path = os.path.join(os.path.dirname(output_path), f"{base_name}_group_pull.json")
        payload = {
            "headers": headers or [],
            "rows": rows_list or [],
        }
        with open(debug_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def rows_to_df(api, headers, rows_list):
    if not rows_list:
        return pd.DataFrame(columns=headers if headers else None)
    parsed = api._parse_rows(rows_list)
    try:
        return pd.DataFrame(parsed, columns=headers if headers else None)
    except Exception:
        return pd.DataFrame(parsed)


def extract_unit_mapping(headers, rows_list):
    """Return dict: unit_id -> unit_name from group report rows."""
    unit_name_to_id = {}
    unit_col_idx = None
    if headers:
        for i, h in enumerate(headers):
            if "group" in str(h).lower() or "unit" in str(h).lower():
                unit_col_idx = i
                break

    for row in rows_list or []:
        if not isinstance(row, dict) or "c" not in row:
            continue
        cells = row.get("c") or []

        unit_name = None
        unit_id = None

        if unit_col_idx is not None and unit_col_idx < len(cells):
            cell = cells[unit_col_idx]
            if isinstance(cell, dict) and "t" in cell:
                unit_name = str(cell.get("t", "")).strip()
            elif isinstance(cell, str):
                unit_name = cell.strip()

        for cell in cells:
            if isinstance(cell, dict) and "u" in cell:
                try:
                    unit_id = int(cell["u"])
                    break
                except Exception:
                    continue

        if unit_name and unit_id:
            unit_name_to_id[unit_name] = unit_id

    return {v: k for k, v in unit_name_to_id.items()}


def find_event_text_column(df):
    if df is None or df.empty:
        return None
    col = find_column(df, ["event text"])
    if col:
        return col
    for c in df.columns:
        lc = str(c).lower()
        if "event" in lc and "time" not in lc:
            return c
    return find_column(df, ["text", "message", "description"])


def build_unit_events(df, unit_name):
    if df is None or df.empty:
        return pd.DataFrame()

    event_time_col = find_column(df, ["event time"])
    time_received_col = find_column(df, ["time received", "received time", "receive time"])
    event_text_col = find_event_text_column(df)
    location_col = find_column(df, ["location", "address", "place"])
    notif_col = find_column(df, ["notification text", "notification", "event type", "type"])

    # Build using the same index as source rows so scalar columns fill correctly.
    out = pd.DataFrame(index=df.index.copy())
    out["Event time"] = df[event_time_col] if event_time_col in df.columns else ""
    out["Time received"] = df[time_received_col] if time_received_col in df.columns else ""
    out["Event text"] = df[event_text_col] if event_text_col in df.columns else ""
    out["Location"] = df[location_col] if location_col in df.columns else ""
    out["Notification text"] = df[notif_col] if notif_col in df.columns else ""
    out["Row Labels"] = str(unit_name).strip() if unit_name is not None else ""

    out["Event time"] = out["Event time"].apply(format_ddmmyyyy_hhmmss)
    out["Time received"] = out["Time received"].apply(format_ddmmyyyy_hhmmss)

    # Clean HTML entities
    out["Event text"] = out["Event text"].apply(lambda v: html.unescape(str(v)) if pd.notna(v) else "")

    # Normalize notification text
    out["Notification text"] = out["Notification text"].fillna("").astype(str).str.strip()
    out.loc[out["Notification text"].isin(["", "Violation", "nan", "None"]), "Notification text"] = (
        out["Event text"].apply(infer_notification_text)
    )
    out["Notification text"] = out["Notification text"].replace("", "UNKNOWN")
    out["Notification text"] = out["Notification text"].astype(str).str.strip().str.upper()

    # Remove excluded events entirely before aggregation/output.
    exclude_mask = out.apply(
        lambda r: should_exclude_event(r.get("Event text", ""), r.get("Notification text", "")),
        axis=1,
    )
    out = out.loc[~exclude_mask].copy()

    return out


def collapse_unit_events(events_df):
    """Keep first occurrence per (unit, event type), count total occurrences."""
    if events_df is None or events_df.empty:
        return events_df

    events_df["__event_dt"] = pd.to_datetime(events_df["Event time"], dayfirst=True, errors="coerce")
    events_df["__key"] = events_df["Row Labels"].astype(str) + "|" + events_df["Notification text"].astype(str)
    events_df["__count"] = events_df.groupby("__key")["__key"].transform("size")

    events_df = events_df.sort_values(["Row Labels", "Notification text", "__event_dt"])
    events_df = events_df.drop_duplicates(subset=["Row Labels", "Notification text"], keep="first")
    events_df["Count of Event time"] = events_df["__count"].astype(int)

    events_df = events_df.drop(columns=["__event_dt", "__key", "__count"], errors="ignore")

    events_df["Action"] = DEFAULT_ACTION
    events_df["Remark"] = DEFAULT_REMARK

    events_df = events_df[[
        "Row Labels",
        "Count of Event time",
        "Event time",
        "Time received",
        "Event text",
        "Location",
        "Notification text",
        "Action",
        "Remark",
    ]]

    return events_df


def build_summary_df(events_df):
    if events_df is None or events_df.empty:
        return pd.DataFrame(columns=["EVENTS", "TOTAL EVENTS", "REMARK"])

    summary = (
        events_df.groupby("Notification text")["Row Labels"]
        .nunique()
        .reset_index()
        .rename(columns={"Notification text": "EVENTS", "Row Labels": "TOTAL EVENTS"})
    )

    summary["EVENTS"] = summary["EVENTS"].fillna("").astype(str).str.strip()
    summary.loc[summary["EVENTS"] == "", "EVENTS"] = "UNKNOWN"
    summary["REMARK"] = summary["EVENTS"].map(SUMMARY_REMARKS).fillna("")

    def _order_key(v):
        return SUMMARY_ORDER.index(v) if v in SUMMARY_ORDER else 999

    summary["__order"] = summary["EVENTS"].apply(_order_key)
    summary = summary.sort_values(["__order", "EVENTS"]).drop(columns=["__order"])

    grand_total = pd.DataFrame([
        {"EVENTS": "Grand Total", "TOTAL EVENTS": int(summary["TOTAL EVENTS"].sum()), "REMARK": ""}
    ])
    summary = pd.concat([summary, grand_total], ignore_index=True)
    return summary


def _copy_row_styles(sample_ws, target_ws, sample_row, target_row, col_count):
    for c in range(1, col_count + 1):
        sample_cell = sample_ws.cell(sample_row, c)
        target_cell = target_ws.cell(target_row, c)
        target_cell._style = copy.copy(sample_cell._style)
        target_cell.number_format = sample_cell.number_format
        target_cell.alignment = copy.copy(sample_cell.alignment)
        target_cell.font = copy.copy(sample_cell.font)
        target_cell.fill = copy.copy(sample_cell.fill)
        target_cell.border = copy.copy(sample_cell.border)
        target_cell.protection = copy.copy(sample_cell.protection)


def _resolve_sample_path(output_path):
    candidates = [
        os.path.join(os.path.dirname(__file__), DEFAULT_SAMPLE_NAME),
        os.path.join(os.path.dirname(output_path), DEFAULT_SAMPLE_NAME),
        os.path.join(os.path.dirname(os.path.dirname(output_path)), DEFAULT_SAMPLE_NAME),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def apply_sample_formatting(output_path, sample_path=None):
    """Apply sheet styles/widths from sample workbook to generated workbook."""
    sample_path = sample_path or _resolve_sample_path(output_path)
    if not sample_path or not os.path.exists(sample_path):
        print("Sample format file not found; skipping formatting copy.")
        return

    sample_wb = openpyxl.load_workbook(sample_path)
    out_wb = openpyxl.load_workbook(output_path)

    for sheet_name in ["EVENTS", "SUMMARY"]:
        if sheet_name not in sample_wb.sheetnames or sheet_name not in out_wb.sheetnames:
            continue

        sample_ws = sample_wb[sheet_name]
        out_ws = out_wb[sheet_name]
        col_count = out_ws.max_column

        # Match sample column widths
        for col_letter, dim in sample_ws.column_dimensions.items():
            if dim.width is not None:
                out_ws.column_dimensions[col_letter].width = dim.width

        # Header style
        _copy_row_styles(sample_ws, out_ws, 1, 1, col_count)

        if out_ws.max_row >= 2:
            # Standard body rows
            for r in range(2, out_ws.max_row + 1):
                _copy_row_styles(sample_ws, out_ws, 2, r, col_count)

        # Summary grand total style for last row
        if sheet_name == "SUMMARY" and out_ws.max_row >= 2:
            sample_last = sample_ws.max_row
            _copy_row_styles(sample_ws, out_ws, sample_last, out_ws.max_row, col_count)

    out_wb.save(output_path)
    sample_wb.close()
    out_wb.close()


def main():
    parser = argparse.ArgumentParser(
        description="Pull GALCO all events report with per-unit per-event aggregation."
    )
    parser.add_argument("output_dir", nargs="?", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("group_name", nargs="?", default=DEFAULT_GROUP)
    parser.add_argument("date_positional", nargs="?", default=None)
    parser.add_argument("--date", dest="date_flag", default=None, help="Exact date (YYYY-MM-DD or DD.MM.YYYY)")
    parser.add_argument(
        "--day",
        dest="day_mode",
        choices=["today", "yesterday"],
        default="today",
        help="Relative day mode when --date is not provided.",
    )
    args = parser.parse_args()

    output_dir = args.output_dir
    group_name = args.group_name
    date_arg = args.date_flag or args.date_positional

    os.makedirs(output_dir, exist_ok=True)

    if date_arg:
        target_date = parse_date_arg(date_arg)
    else:
        target_date = get_system_today_date() if args.day_mode == "today" else (get_system_today_date() - timedelta(days=1))
    today_tz = get_system_today_date()
    interval_from, interval_to = get_day_interval_tz(
        target_date,
        until_now=(target_date == today_tz),
    )
    date_str = target_date.strftime("%d.%m.%Y")

    output_path = os.path.join(output_dir, f"{group_name}_EVENTS_{date_str}.xlsx")

    api = WialonAPI()
    if not api.login():
        print("Failed to login to Wialon")
        return 1

    try:
        group_id = api.find_group_id(group_name)
        if not group_id:
            print(f"Group not found: {group_name}")
            return 1

        g_headers, g_rows, _ = exec_report_raw(
            api, group_id, GROUP_TEMPLATE_ID, interval_from, interval_to, output_path
        )
        if g_headers is None:
            print("Failed to fetch group report")
            return 1
        save_group_debug_json(output_path, g_headers, g_rows)

        unit_map = extract_unit_mapping(g_headers, g_rows)
        unit_ids = list(unit_map.keys())
        if not unit_ids:
            print("No units found in group report for this date.")
            return 0

        all_events = []
        print(f"Pulling single-unit report for {len(unit_ids)} units...")
        for idx, unit_id in enumerate(unit_ids, 1):
            unit_name = unit_map.get(unit_id, f"UNIT_{unit_id}")
            print(f"  [{idx}/{len(unit_ids)}] {unit_name} ({unit_id})")

            headers, rows_list, _ = exec_report_raw(
                api, unit_id, UNIT_TEMPLATE_ID, interval_from, interval_to, output_path
            )
            if headers is None:
                continue
            save_unit_debug_json(output_path, unit_id, unit_name, headers, rows_list)
            df = rows_to_df(api, headers, rows_list)
            if df is None or df.empty:
                continue
            unit_events = build_unit_events(df, unit_name)
            if not unit_events.empty:
                all_events.append(unit_events)
            time.sleep(0.25)

        if not all_events:
            print("No events found in single-unit reports.")
            return 0

        events_df = pd.concat(all_events, ignore_index=True)
        events_df = collapse_unit_events(events_df)
        summary_df = build_summary_df(events_df)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            events_df.to_excel(writer, sheet_name="EVENTS", index=False)
            summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)
        apply_sample_formatting(output_path)

        print(f"Report saved: {output_path}")
        return 0

    finally:
        api.logout()


if __name__ == "__main__":
    raise SystemExit(main())
